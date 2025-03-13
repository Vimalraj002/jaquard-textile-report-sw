import tkinter as tk
from tkinter import ttk
from tkcalendar import DateEntry
from openpyxl import load_workbook
import openpyxl
from datetime import datetime

# Load dropdown values from separate Excel files
def load_excel_data(filename):
    workbook = load_workbook(filename)
    sheet = workbook.active
    return [cell.value for cell in sheet['A'] if cell.value]

# Load customer, weaver, and lacing lists
customer_list = load_excel_data('Customers.xlsx')
weaver_list = load_excel_data('Weaver.xlsx')
lacing_list = load_excel_data('Lacing.xlsx')
item_list = load_excel_data('Items.xlsx')

# Set up Excel file for data entry
data_file = 'data_entry.xlsx'
try:
    wb = openpyxl.load_workbook(data_file)
except FileNotFoundError:
    wb = openpyxl.Workbook()
    sheet = wb.active
    sheet.title = "Data"
    headers = ["Date", "Customer", "Item", "Design Number", "Weaver", "Lacing", "Cards"]
    sheet.append(headers)
    wb.save(data_file)

# Function to filter dropdown values dynamically
def filter_dropdown(event, combobox, options):
    input_text = combobox.get().lower()  # Get the text in lowercase to make it case-insensitive
    if not input_text:  # Reset to full list if input is empty
        combobox['values'] = sorted(options)  # Sort the list alphabetically
    else:
        # Filter options: match only words that start with the typed input
        filtered_options = [option for option in options if option.lower().startswith(input_text)]
        combobox['values'] = sorted(filtered_options) if filtered_options else ["No match found"]
        
    # Manually open the dropdown after updating values
    combobox.event_generate('<Down>')  # Open the dropdown



# Function to reset dropdown values when clearing
def reset_dropdown_values():
    customer_dropdown['values'] = customer_list
    weaver_dropdown['values'] = weaver_list
    lacing_dropdown['values'] = lacing_list
    item_dropdown['values'] = lacing_list

# Function to clear form fields
def clear_fields():
    today_date = datetime.today().date()
    date_entry.set_date(today_date)
    customer_var.set('')
    item_var.set('')
    weaver_var.set('')
    lacing_var.set('')
    entry_design_number.delete(0, tk.END)
    entry_cards.delete(0, tk.END)
    reset_dropdown_values()
    global editing_record
    editing_record = None  # Clear editing state
    status_label.config(text="Fields cleared.", fg="blue")

def save_data():
    global editing_record
    date = date_entry.get()
    customer = customer_var.get()
    item = item_var.get()
    design_number = entry_design_number.get()
    weaver = weaver_var.get()
    lacing = lacing_var.get()
    cards = entry_cards.get()

    if not (date and customer and item and design_number and weaver and lacing and cards):
        status_label.config(text="Please fill all fields.", fg="red")
        return

    if editing_record is None:
        # Add new data
        wb = openpyxl.load_workbook(data_file)
        sheet = wb['Data']
        sheet.append([date, customer, item, design_number, weaver, lacing, cards])
        wb.save(data_file)
        load_data_to_table()
        status_label.config(text="Data Saved Successfully!", fg="green")
    else:
        # Update existing data
        wb = openpyxl.load_workbook(data_file)
        sheet = wb['Data']
        selected_values = tree.item(editing_record, 'values')

        # Update the matching row in Excel
        for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row):
            if tuple(cell.value for cell in row) == selected_values:
                row[0].value = date
                row[1].value = customer
                row[2].value = item
                row[3].value = design_number
                row[4].value = weaver
                row[5].value = lacing
                row[6].value = cards
                break

        wb.save(data_file)

        # Update Treeview
        tree.item(editing_record, values=(date, customer, item, design_number, weaver, lacing, cards))
        status_label.config(text="Record Updated Successfully!", fg="green")

    # Delay clearing the fields to allow the success message to be visible
    root.after(2000, clear_fields)  # 2000 ms = 2 seconds



# Load data from Excel into Treeview
def load_data_to_table():
    for row in tree.get_children():
        tree.delete(row)
    wb = load_workbook(data_file)
    sheet = wb['Data']
    for row in sheet.iter_rows(min_row=2, values_only=True):
        tree.insert('', 'end', values=row)

# Function to delete selected data
def delete_data():
    selected_item = tree.selection()
    if not selected_item:
        status_label.config(text="No record selected to delete.", fg="red")
        return

    wb = load_workbook(data_file)
    sheet = wb['Data']
    selected_values = tree.item(selected_item, 'values')
    
    # Find and delete the matching row in Excel
    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row):
        if tuple(cell.value for cell in row) == selected_values:
            sheet.delete_rows(row[0].row)
            break

    wb.save(data_file)
    tree.delete(selected_item)
    status_label.config(text="Record deleted successfully.", fg="green")

# Function to populate form with selected data for editing
def edit_data():
    global editing_record
    selected_item = tree.selection()
    if not selected_item:
        status_label.config(text="No record selected to edit.", fg="red")
        return

    selected_values = tree.item(selected_item, 'values')

    # Populate form fields with selected record's data
    date_entry.set_date(datetime.strptime(selected_values[0], '%d/%m/%Y'))
    customer_var.set(selected_values[1])
    item_var.set(selected_values[2])
    entry_design_number.delete(0, tk.END)
    entry_design_number.insert(0, selected_values[3])
    weaver_var.set(selected_values[4])
    lacing_var.set(selected_values[5])
    entry_cards.delete(0, tk.END)
    entry_cards.insert(0, selected_values[6])

    editing_record = selected_item
    status_label.config(text="Record loaded for editing. Make changes and press Submit to save.", fg="blue")

# Function to apply filters
def apply_filters():
    from_date = filter_from_date.get_date()
    to_date = filter_to_date.get_date()
    customer_filter = filter_customer_var.get()

    # Clear the table
    for row in tree.get_children():
        tree.delete(row)

    # Load workbook and sheet
    wb = load_workbook(data_file)
    sheet = wb['Data']
    matching_records = 0

    # Iterate through rows and apply filters
    for row in sheet.iter_rows(min_row=2, values_only=True):
        row_date = datetime.strptime(row[0], "%d/%m/%Y").date()  # Convert string to date
        if from_date <= row_date <= to_date and (not customer_filter or customer_filter == row[1]):
            tree.insert("", "end", values=row)
            matching_records += 1

    # Update status
    if matching_records > 0:
        status_label.config(
            text=f"Filters applied successfully. {matching_records} record(s) found.", fg="green"
        )
    else:
        status_label.config(text="No matching records found.", fg="red")

    from_date = filter_from_date.get_date()
    to_date = filter_to_date.get_date()
    customer = filter_customer_var.get()
    load_data_to_table(from_date, to_date, customer)

# Modify the load_data_to_table function to handle filters
def load_data_to_table(from_date=None, to_date=None, customer=None):
    for row in tree.get_children():
        tree.delete(row)
    wb = load_workbook(data_file)
    sheet = wb['Data']
    for row in sheet.iter_rows(min_row=2, values_only=True):
        row_date = datetime.strptime(row[0], '%d/%m/%Y').date()
        if (
            (not from_date or row_date >= from_date) and
            (not to_date or row_date <= to_date) and
            (not customer or customer.lower() in row[1].lower())
        ):
            tree.insert('', 'end', values=row)




# GUI Setup
root = tk.Tk()
root.title("Data Entry")
root.geometry("900x800")
root.configure(bg="#f2f2f2")

form_frame = tk.Frame(root, bg="#ffffff", bd=2, relief="groove")
form_frame.pack(pady=10, padx=10, fill='x')

tk.Label(form_frame, text="Date", bg="#ffffff").grid(row=0, column=0, padx=10, pady=5)
date_entry = DateEntry(form_frame, width=12, background='darkblue', foreground='white', borderwidth=2, date_pattern='dd/mm/yyyy')
date_entry.grid(row=0, column=1, padx=10, pady=5)

tk.Label(form_frame, text="Customer", bg="#ffffff").grid(row=1, column=0, padx=10, pady=5)
customer_var = tk.StringVar()
customer_dropdown = ttk.Combobox(form_frame, textvariable=customer_var)
customer_dropdown['values'] = customer_list
customer_dropdown.grid(row=1, column=1, padx=10, pady=5)
customer_dropdown.bind('<KeyRelease>', lambda e: filter_dropdown(e, customer_dropdown, customer_list))


tk.Label(form_frame, text="Item", bg="#ffffff").grid(row=2, column=0, padx=10, pady=5)
item_var = tk.StringVar()
item_dropdown = ttk.Combobox(form_frame, textvariable=item_var)
item_dropdown['values'] = item_list
item_dropdown.grid(row=2, column=1, padx=10, pady=5)
item_dropdown.bind('<KeyRelease>', lambda e: filter_dropdown(e, item_dropdown, item_list))

tk.Label(form_frame, text="Design Number", bg="#ffffff").grid(row=3, column=0, padx=10, pady=5)
entry_design_number = tk.Entry(form_frame)
entry_design_number.grid(row=3, column=1, padx=10, pady=5)

tk.Label(form_frame, text="Weaver", bg="#ffffff").grid(row=4, column=0, padx=10, pady=5)
weaver_var = tk.StringVar()
weaver_dropdown = ttk.Combobox(form_frame, textvariable=weaver_var)
weaver_dropdown['values'] = weaver_list
weaver_dropdown.grid(row=4, column=1, padx=10, pady=5)
weaver_dropdown.bind('<KeyRelease>', lambda e: filter_dropdown(e, weaver_dropdown, weaver_list))

tk.Label(form_frame, text="Lacing", bg="#ffffff").grid(row=5, column=0, padx=10, pady=5)
lacing_var = tk.StringVar()
lacing_dropdown = ttk.Combobox(form_frame, textvariable=lacing_var)
lacing_dropdown['values'] = lacing_list
lacing_dropdown.grid(row=5, column=1, padx=10, pady=5)
lacing_dropdown.bind('<KeyRelease>', lambda e: filter_dropdown(e, lacing_dropdown, lacing_list))

tk.Label(form_frame, text="Cards", bg="#ffffff").grid(row=6, column=0, padx=10, pady=5)
entry_cards = tk.Entry(form_frame)
entry_cards.grid(row=6, column=1, padx=10, pady=5)

# Button frame
button_frame = tk.Frame(form_frame, bg="#ffffff")
button_frame.grid(row=7, column=0, columnspan=4, pady=10, sticky="w")

# Action buttons
submit_button = tk.Button(button_frame, text="Submit", command=save_data, bg="#4CAF50", fg="white", width=10)
submit_button.pack(side="left", padx=2)

edit_button = tk.Button(button_frame, text="Edit", command=edit_data, bg="#FFC107", fg="black", width=10)
edit_button.pack(side="left", padx=2)

delete_button = tk.Button(button_frame, text="Delete", command=delete_data, bg="#f44336", fg="white", width=10)
delete_button.pack(side="left", padx=2)

clear_button = tk.Button(button_frame, text="Clear", command=clear_fields, bg="#00bcd4", fg="white", width=10)
clear_button.pack(side="left", padx=2)



status_label = tk.Label(form_frame, text="", bg="#ffffff")
status_label.grid(row=8, column=0, columnspan=4, pady=5)

tree_frame = tk.Frame(root)
tree_frame.pack(padx=10, pady=10, fill='x')

# Filter controls
tk.Label(form_frame, text="From Date", bg="#ffffff").grid(row=0, column=3, padx=10, pady=5, sticky="w")
filter_from_date = DateEntry(form_frame, width=12, background='darkblue', foreground='white', borderwidth=2, date_pattern='dd/mm/yyyy')
filter_from_date.grid(row=0, column=4, padx=10, pady=5, sticky="w")

tk.Label(form_frame, text="To Date", bg="#ffffff").grid(row=1, column=3, padx=10, pady=5, sticky="w")
filter_to_date = DateEntry(form_frame, width=12, background='darkblue', foreground='white', borderwidth=2, date_pattern='dd/mm/yyyy')
filter_to_date.grid(row=1, column=4, padx=10, pady=5, sticky="w")

tk.Label(form_frame, text="Customer", bg="#ffffff").grid(row=2, column=3, padx=10, pady=5, sticky="w")
filter_customer_var = tk.StringVar()
filter_customer_dropdown = ttk.Combobox(form_frame, textvariable=filter_customer_var)
filter_customer_dropdown['values'] = customer_list
filter_customer_dropdown.grid(row=2, column=4, padx=10, pady=5, sticky="w")

apply_filter_button = tk.Button(form_frame, text="Apply Filters", command=apply_filters, bg="#4CAF50", fg="white")
apply_filter_button.grid(row=3, column=3, padx=10, pady=10, sticky="w")

clear_filter_button = tk.Button(
    form_frame,
    text="Clear Filters",
    command=lambda: [
        load_data_to_table(),
        status_label.config(text="Filters cleared. Showing all records.", fg="blue")
    ],
    bg="#00bcd4",
    fg="white"
)
clear_filter_button.grid(row=3, column=4, padx=10, pady=10, sticky="w")

form_frame.grid_columnconfigure(0, weight=1)
form_frame.grid_columnconfigure(1, weight=1)
form_frame.grid_columnconfigure(2, weight=1)
form_frame.grid_columnconfigure(3, weight=1)
form_frame.grid_columnconfigure(4, weight=1)



columns = ["Date", "Customer", "Item", "Design Number", "Weaver", "Lacing", "Cards"]
tree = ttk.Treeview(tree_frame, columns=columns, show='headings')
tree.pack(fill="both", expand=True)
for col in columns:
    tree.heading(col, text=col)
    tree.column(col, width=120)

load_data_to_table()

editing_record = None  # Global variable to track the record being edited

root.mainloop()