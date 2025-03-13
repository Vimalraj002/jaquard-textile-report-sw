import tkinter as tk
from tkinter import ttk, messagebox
import os
# import subprocess
from openpyxl import Workbook, load_workbook

# File to track payment statuses
payment_status_file = 'customer_payment_status.xlsx'

# Ensure the payment status file exists
if not os.path.exists(payment_status_file):
    wb = Workbook()
    sheet = wb.active
    sheet.title = "Payment Status"
    sheet.append(["Invoice Number", "Report Name", "Payment Status"])
    wb.save(payment_status_file)

# Function to load reports into the Treeview
def load_reports(report_tree):
    report_tree.delete(*report_tree.get_children())
    reports_dir = "Customer Reports"
    if not os.path.exists(reports_dir):
        os.makedirs(reports_dir)
    reports = [file for file in os.listdir(reports_dir) if file.endswith('.pdf')]

    # Load payment statuses
    wb = load_workbook(payment_status_file)
    sheet = wb['Payment Status']
    status_dict = {row[1].value: (row[0].value, row[2].value) for row in sheet.iter_rows(min_row=2)}

    for report in reports:
        invoice_number, status = status_dict.get(report, ("", "Unpaid"))
        report_tree.insert('', 'end', values=(invoice_number, report, status))

# Function to view a PDF report
def view_pdf(report_name):
    try:
        reports_dir = "Customer Reports"
        file_path = os.path.join(reports_dir, report_name)
        if os.name == 'nt':  # For Windows
            os.startfile(file_path)
        elif os.name == 'posix':  # For macOS/Linux
            subprocess.call(('open', file_path))
    except Exception as e:
        messagebox.showerror("Error", f"Unable to open PDF: {e}")

def mark_payment_status(report_tree, status):
    selected_item = report_tree.selection()
        
    if not selected_item:
        messagebox.showwarning("No Selection", "Please select a report to mark.")
        return

    # Get the report name from Treeview
    report_name = report_tree.item(selected_item)['values'][1]
    if not report_name:
        messagebox.showerror("Error", "Could not retrieve report name.")
        return

    # Open the workbook and update the status
    wb = load_workbook(payment_status_file)
    sheet = wb['Payment Status']

    found = False  # Track if the report is found in the Excel file
    for row in sheet.iter_rows(min_row=2, max_col=3):
        if row[1].value and row[1].value.strip() == report_name.strip():
            row[2].value = status
            found = True
            break

    if not found:
        # If not found, add a new entry
        sheet.append([None, report_name, status])

    # Save the workbook and reload the Treeview
    wb.save(payment_status_file)
    load_reports(report_tree)

# Main application window
root = tk.Tk()
root.title("Customer Payments")
root.geometry("900x800")

# Treeview to display reports
columns = ["Invoice Number", "Report Name", "Payment Status"]
report_tree = ttk.Treeview(root, columns=columns, show='headings')
for col in columns:
    report_tree.heading(col, text=col)
    report_tree.column(col, width=300 if col == "Report Name" else 150)
report_tree.pack(fill='both', expand=True, padx=10, pady=10)

# Load reports
load_reports(report_tree)

# Buttons for actions
button_frame = tk.Frame(root)
button_frame.pack(pady=10)

view_button = tk.Button(button_frame, text="View Report", command=lambda: view_pdf(report_tree.item(report_tree.selection())['values'][1]), bg="#2196F3", fg="white")
view_button.grid(row=0, column=0, padx=10)

paid_button = tk.Button(button_frame, text="Mark as Paid", command=lambda: mark_payment_status(report_tree, "Paid"), bg="#4CAF50", fg="white")
paid_button.grid(row=0, column=1, padx=10)

unpaid_button = tk.Button(button_frame, text="Mark as Unpaid", command=lambda: mark_payment_status(report_tree, "Unpaid"), bg="#f44336", fg="white")
unpaid_button.grid(row=0, column=2, padx=10)

root.mainloop()
