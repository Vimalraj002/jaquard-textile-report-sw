import tkinter as tk
from tkinter import ttk
from tkcalendar import DateEntry
from openpyxl import Workbook, load_workbook
from fpdf import FPDF
from datetime import datetime
import os
from collections import defaultdict


# Load customer list for filtering
def load_excel_data(filename):
    if not os.path.exists(filename):
        return []
    workbook = load_workbook(filename)
    sheet = workbook.active
    return [cell.value for cell in sheet['A'] if cell.value]


# Load customer list
customer_list = load_excel_data('Customers.xlsx')

# Set up Excel file for data entry
data_file = 'data_entry.xlsx'
if not os.path.exists(data_file):
    raise FileNotFoundError(f"Data file '{data_file}' not found. Please ensure it exists.")


# Filter data function
def filter_data(sheet, start_date, end_date, selected_customer):
    filtered_data = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        date_str = row[0]
        customer = row[1]

        try:
            date = datetime.strptime(date_str, "%d/%m/%Y").date() if isinstance(date_str, str) else date_str
        except ValueError:
            continue

        if start_date <= date <= end_date and (not selected_customer or customer == selected_customer):
            filtered_data.append(row)
    return filtered_data


# Clear existing data in the TreeView table
def clear_table():
    for row in tree.get_children():
        tree.delete(row)


# Function to display data in the TreeView table
def display_data_in_table(data):
    clear_table()
    for row in data:
        tree.insert("", "end", values=row)


# Adjust TreeView column widths dynamically
def adjust_tree_columns(tree, data):
    for i, col in enumerate(columns):
        max_width = max(len(str(row[i])) for row in data) if data else 10
        tree.column(col, width=max_width * 10)

from openpyxl import Workbook, load_workbook
import os

def save_payment_status(invoice_number, report_name, payment_status='Unpaid'):
    # File path for the payment status file
    payment_status_file = 'customer_payment_status.xlsx'
    
    # Check if the file exists; if not, create it with appropriate headers
    if not os.path.exists(payment_status_file):
        wb = Workbook()
        sheet = wb.active
        sheet.title = "Payment Status"
        # Add headers
        sheet.append(['Invoice Number', 'Report Name', 'Payment Status'])
        wb.save(payment_status_file)
    
    # Load the existing file
    wb = load_workbook(payment_status_file)
    sheet = wb.active

    # Append the new record
    try:
        sheet.append([
            invoice_number,
            report_name,
            payment_status
        ])
        wb.save(payment_status_file)
        print("Payment status saved successfully.")
    except Exception as e:
        print(f"Error saving payment status: {e}")
        
# Function to generate the report
# Update the generate_report function to call save_payment_status
def generate_report():
    start_date = start_date_entry.get_date()
    end_date = end_date_entry.get_date()
    selected_customer = customer_var.get()
    invoice_number = invoice_number_entry.get()

    wb = load_workbook(data_file)
    sheet = wb['Data']

    filtered_data = filter_data(sheet, start_date, end_date, selected_customer)
    if not filtered_data:
        result_label.config(text="No data found for the selected filter", fg="red")
        clear_table()
        return

    result_label.config(text=f"Found {len(filtered_data)} records.", fg="green")
    display_data_in_table(filtered_data)
    adjust_tree_columns(tree, filtered_data)

# Function to generate the PDF report
def generate_pdf_report(filtered_data, start_date, end_date, selected_customer, invoice_number):
    if not filtered_data:
        result_label.config(text="No data to generate report.", fg="red")
        return

    # Group data by item
    grouped_data = defaultdict(list)
    for row in filtered_data:
        item = row[2]  # Assuming "Item" is in the 3rd column
        grouped_data[item].append(row)

    pdf = FPDF()
    pdf.add_page()
    pdf.set_font("Arial", size=10)

    pdf.set_font("Arial", size=10,style='B')
    pdf.cell(50, 10, txt="Customer Name:", border=0)
    pdf.cell(100, 10, txt=selected_customer or "All Customers", border=0)
    pdf.ln(5)

    pdf.cell(50, 10, txt="Start Date:", border=0)
    pdf.cell(100, 10, txt=start_date.strftime('%d/%m/%Y'), border=0)
    pdf.ln(5)

    pdf.cell(50, 10, txt="End Date:", border=0)
    pdf.cell(100, 10, txt=end_date.strftime('%d/%m/%Y'), border=0)
    pdf.ln(5)

    # Add the Invoice Number
    pdf.cell(50, 10, txt="Invoice Number:", border=0)
    pdf.cell(100, 10, txt=invoice_number or "Not Provided", border=0)
    pdf.ln(10)

    # Table Header with light gray background and bold font
    header = ['Date', 'Item', 'Design No', 'Weaver Name', 'Cards']
    pdf.set_fill_color(200, 200, 200)  # Light gray background
    pdf.set_font("Arial", style='B', size=10)  # Bold font

    for header_text in header:
        pdf.cell(38, 7, header_text, border=1, align='C', fill=True)  # Set fill=True for background color
    pdf.ln()  # Move to the next line


    # Table Rows and Subtotals
    grand_total = 0
    for item, rows in grouped_data.items():
        item_total = 0
        for row in rows:
             # Reset to normal font and text color for regular rows
            pdf.set_font("Arial", size=10)
            pdf.set_text_color(0, 0, 0)  # Black text
            pdf.cell(38, 5, row[0].strftime('%d/%m/%Y') if isinstance(row[0], datetime) else str(row[0]), border=0, align='L')
            pdf.cell(38, 5, str(row[2]), border=0, align='L')  # Item
            pdf.cell(38, 5, str(row[3]), border=0, align='L')  # Design No
            pdf.cell(40, 5, str(row[4]), border=0, align='L')  # Weaver Name
            pdf.cell(38, 5, str(row[6]), border=0, align='C')  # Cards (converted to string)
            pdf.ln()
            item_total += row[6] if isinstance(row[6], (int, float)) else 0

        # Set the background color for the subtotal row (e.g., light gray)
        pdf.set_fill_color(230, 230, 230)  # Light gray background

        # Set the font to bold for the subtotal row
        pdf.set_font("Arial", size=10)

        # Set the text color to red for the subtotal
        pdf.set_text_color(255, 0, 0)  # Red text

        # Subtotal for the item
        pdf.cell(154, 5, f"{item} (Total)", border=0, align='R', fill=True)  # 'fill=True' to use the background color
        pdf.cell(38, 5, str(item_total), border=0, align='C', fill=True)  # Converted to string, with background color
        pdf.ln()

        # Reset font and text color for subsequent rows
        pdf.set_font("Arial", size=10)
        pdf.set_text_color(0, 0, 0)  # Reset text color to black

        # Add the item total to the grand total
        grand_total += item_total

    # Save PDF
    report_dir = "Customer Reports"
    os.makedirs(report_dir, exist_ok=True)
    # Replace any invalid characters in the customer name
    sanitized_customer = selected_customer.replace("/", "-").replace("\\", "-")
    # Use dd-mm-yyyy format for the dates in the filename
    report_filename = f"{sanitized_customer}({invoice_number}) - {start_date.strftime('%d-%m-%Y')} to {end_date.strftime('%d-%m-%Y')}.pdf"
    pdf_output_path = os.path.join(report_dir, report_filename)
    pdf.output(pdf_output_path)
    result_label.config(text=f"Report generated: {pdf_output_path}", fg="green")

    # Save the invoice number and customer info to the payment_status.xlsx file
    save_payment_status(invoice_number, report_filename)


# GUI Setup
root = tk.Tk()
root.title("Customer Report")
root.geometry("800x600")

# Report filters
report_frame = tk.Frame(root)
report_frame.pack(pady=10)

start_date_label = tk.Label(report_frame, text="Start Date")
start_date_label.grid(row=0, column=0, padx=10, pady=5)

start_date_entry = DateEntry(report_frame, width=12, background='darkblue', foreground='white', borderwidth=2, date_pattern='dd/mm/yyyy')
start_date_entry.grid(row=0, column=1, padx=10, pady=5)

end_date_label = tk.Label(report_frame, text="End Date")
end_date_label.grid(row=1, column=0, padx=10, pady=5)

end_date_entry = DateEntry(report_frame, width=12, background='darkblue', foreground='white', borderwidth=2, date_pattern='dd/mm/yyyy')
end_date_entry.grid(row=1, column=1, padx=10, pady=5)

customer_label = tk.Label(report_frame, text="Customer")
customer_label.grid(row=2, column=0, padx=10, pady=5)

customer_var = tk.StringVar()
customer_dropdown = ttk.Combobox(report_frame, textvariable=customer_var, values=customer_list)
customer_dropdown.grid(row=2, column=1, padx=10, pady=5)

invoice_number_label = tk.Label(report_frame, text="Invoice Number")
invoice_number_label.grid(row=3, column=0, padx=10, pady=5)

invoice_number_entry = tk.Entry(report_frame, width=20)
invoice_number_entry.grid(row=3, column=1, padx=10, pady=5)

generate_button = tk.Button(report_frame, text="Show Entries", command=generate_report, bg="#4CAF50", fg="white")
generate_button.grid(row=4, column=0, columnspan=2, pady=20)

result_label = tk.Label(report_frame, text="", fg="green")
result_label.grid(row=5, column=0, columnspan=2)

# TreeView for displaying filtered data
tree_frame = tk.Frame(root)
tree_frame.pack(fill="both", expand=True, pady=10)

columns = ['Date', 'Customer', 'Item', 'Design Number', 'Weaver', 'Cards']
tree = ttk.Treeview(tree_frame, columns=columns, show='headings', height=10)
tree.pack(fill="both", expand=True)

for col in columns:
    tree.heading(col, text=col)
    tree.column(col, anchor="center", width=120)

# Button to generate PDF report from the table
pdf_button = tk.Button(root, text="Download", command=lambda: generate_pdf_report(
    [tree.item(row)["values"] for row in tree.get_children()],
    start_date_entry.get_date(),
    end_date_entry.get_date(),
    customer_var.get(),
    invoice_number_entry.get()
), bg="#FF5733", fg="white")
pdf_button.pack(pady=10)

root.mainloop()